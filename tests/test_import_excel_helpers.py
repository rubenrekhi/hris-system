"""
Helper functions for creating test Excel files.
"""

import io
from datetime import date, datetime
import openpyxl
from openpyxl import Workbook


def create_excel_file(data: list[dict]) -> io.BytesIO:
    """
    Create an Excel file from a list of dictionaries.

    Args:
        data: List of dicts where keys are column headers

    Returns:
        BytesIO object containing the Excel file
    """
    if not data:
        wb = Workbook()
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes

    wb = Workbook()
    ws = wb.active

    # Write headers (from first dict)
    headers = list(data[0].keys())
    ws.append(headers)

    # Write data rows
    for row_dict in data:
        row = [row_dict.get(h, "") for h in headers]
        ws.append(row)

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes


def create_excel_with_types(data: list[dict], type_overrides: dict = None) -> io.BytesIO:
    """
    Create an Excel file with specific data types.

    Args:
        data: List of dicts with data
        type_overrides: Dict mapping column names to type conversion functions

    Returns:
        BytesIO object containing the Excel file
    """
    if not data:
        return create_excel_file(data)

    wb = Workbook()
    ws = wb.active

    headers = list(data[0].keys())
    ws.append(headers)

    type_overrides = type_overrides or {}

    for row_dict in data:
        row = []
        for header in headers:
            value = row_dict.get(header, "")

            # Apply type override if specified
            if header in type_overrides:
                converter = type_overrides[header]
                value = converter(value) if value != "" else ""

            row.append(value)

        ws.append(row)

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes
