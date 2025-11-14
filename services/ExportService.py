"""
ExportService.py
----------------
Business logic for exporting employee data in various formats.
"""

from __future__ import annotations
from typing import List, Dict, Any
from uuid import UUID
from datetime import date
from io import BytesIO
import csv
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select
from models.EmployeeModel import Employee, EmployeeStatus
from schemas.ExportSchemas import ExportFilterSchema

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# PDF imports
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas


class ExportService:
    """Service for exporting employee data."""

    def __init__(self, db: Session):
        self.db = db

    # ============================================================================
    # Private Helper Methods
    # ============================================================================

    def _apply_filters(self, filters: ExportFilterSchema) -> List[Employee]:
        """
        Apply filters to employee query and return all matching employees.

        Args:
            filters: ExportFilterSchema with filter criteria

        Returns:
            List of Employee objects matching the filters
        """
        # Build filters list
        filter_conditions = []

        if filters.department_id:
            filter_conditions.append(Employee.department_id == filters.department_id)

        if filters.team_id:
            filter_conditions.append(Employee.team_id == filters.team_id)

        if filters.status:
            filter_conditions.append(Employee.status == filters.status)

        if filters.hired_from:
            filter_conditions.append(Employee.hired_on >= filters.hired_from)

        if filters.hired_to:
            filter_conditions.append(Employee.hired_on <= filters.hired_to)

        # Build query with eager loading for relationships
        query = (
            select(Employee)
            .options(
                joinedload(Employee.department),
                joinedload(Employee.team),
                joinedload(Employee.manager),
            )
            .order_by(Employee.name)
        )

        # Apply filters
        if filter_conditions:
            query = query.where(*filter_conditions)

        # Execute query
        result = self.db.execute(query)
        employees = result.scalars().unique().all()

        return list(employees)

    def _build_org_tree(self, employees: List[Employee]) -> List[Dict[str, Any]]:
        """
        Build a hierarchical organization tree from a flat list of employees.

        Returns a list of dicts with employee data plus 'level' for indentation.
        The list is ordered depth-first so managers come before their reports.

        Args:
            employees: List of Employee objects

        Returns:
            List of dicts with employee data plus 'level' field for hierarchy
        """
        # Create a map of employee_id -> employee
        emp_map = {emp.id: emp for emp in employees}

        # Find root employees (those with no manager or manager not in filtered set)
        roots = [
            emp for emp in employees
            if emp.manager_id is None or emp.manager_id not in emp_map
        ]

        # Recursively build the tree
        result = []

        def add_employee_and_reports(emp: Employee, level: int):
            """Recursively add employee and their reports to result."""
            # Add current employee
            result.append({
                'employee': emp,
                'level': level,
            })

            # Find and add direct reports
            reports = [e for e in employees if e.manager_id == emp.id]
            reports.sort(key=lambda e: e.name)  # Sort reports alphabetically

            for report in reports:
                add_employee_and_reports(report, level + 1)

        # Process each root
        for root in sorted(roots, key=lambda e: e.name):
            add_employee_and_reports(root, 0)

        return result

    def _employee_to_dict(self, employee: Employee) -> Dict[str, Any]:
        """
        Convert an Employee object to a dictionary for export.

        Args:
            employee: Employee object

        Returns:
            Dictionary with employee data
        """
        return {
            'Name': employee.name,
            'Title': employee.title or '',
            'Email': employee.email,
            'Status': employee.status.value if employee.status else '',
            'Department': employee.department.name if employee.department else '',
            'Team': employee.team.name if employee.team else '',
            'Manager': employee.manager.name if employee.manager else '',
            'Hired On': employee.hired_on.isoformat() if employee.hired_on else '',
            'Salary': employee.salary if employee.salary is not None else '',
        }

    # ============================================================================
    # Directory Export Methods (Flat List)
    # ============================================================================

    def export_directory_csv(self, filters: ExportFilterSchema) -> BytesIO:
        """
        Export employee directory as CSV.

        Args:
            filters: Export filter criteria

        Returns:
            BytesIO buffer containing CSV data
        """
        employees = self._apply_filters(filters)

        # Create CSV in memory
        output = BytesIO()
        # Use text mode wrapper
        text_output = output

        # Write to string first, then encode
        import io
        string_output = io.StringIO()

        fieldnames = ['Name', 'Title', 'Email', 'Status', 'Department', 'Team', 'Manager', 'Hired On', 'Salary']
        writer = csv.DictWriter(string_output, fieldnames=fieldnames)

        writer.writeheader()
        for emp in employees:
            writer.writerow(self._employee_to_dict(emp))

        # Get the string value and encode it
        output.write(string_output.getvalue().encode('utf-8'))
        output.seek(0)

        return output

    def export_directory_excel(self, filters: ExportFilterSchema) -> BytesIO:
        """
        Export employee directory as Excel.

        Args:
            filters: Export filter criteria

        Returns:
            BytesIO buffer containing Excel data
        """
        employees = self._apply_filters(filters)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Directory"

        # Define headers
        headers = ['Name', 'Title', 'Email', 'Status', 'Department', 'Team', 'Manager', 'Hired On', 'Salary']

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, emp in enumerate(employees, 2):
            emp_dict = self._employee_to_dict(emp)
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=emp_dict[header])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def export_directory_pdf(self, filters: ExportFilterSchema) -> BytesIO:
        """
        Export employee directory as PDF table.

        Args:
            filters: Export filter criteria

        Returns:
            BytesIO buffer containing PDF data
        """
        employees = self._apply_filters(filters)

        # Create PDF in memory
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter))

        # Container for PDF elements
        elements = []

        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
        )
        title = Paragraph("Employee Directory", title_style)
        elements.append(title)

        # Prepare table data
        headers = ['Name', 'Title', 'Email', 'Status', 'Dept', 'Team', 'Manager', 'Hired', 'Salary']
        data = [headers]

        for emp in employees:
            emp_dict = self._employee_to_dict(emp)
            row = [
                emp_dict['Name'],
                emp_dict['Title'][:20] if emp_dict['Title'] else '',  # Truncate for space
                emp_dict['Email'],
                emp_dict['Status'],
                emp_dict['Department'][:15] if emp_dict['Department'] else '',
                emp_dict['Team'][:15] if emp_dict['Team'] else '',
                emp_dict['Manager'][:20] if emp_dict['Manager'] else '',
                emp_dict['Hired On'],
                str(emp_dict['Salary']) if emp_dict['Salary'] else '',
            ]
            data.append(row)

        # Create table
        table = Table(data, repeatRows=1)

        # Style table
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        output.seek(0)

        return output

    # ============================================================================
    # Org Chart Export Methods (Hierarchical)
    # ============================================================================

    def export_org_chart_csv(self, filters: ExportFilterSchema) -> BytesIO:
        """
        Export organizational chart as CSV with indentation.

        Args:
            filters: Export filter criteria

        Returns:
            BytesIO buffer containing CSV data
        """
        employees = self._apply_filters(filters)
        org_tree = self._build_org_tree(employees)

        # Create CSV in memory
        import io
        string_output = io.StringIO()

        fieldnames = ['Level', 'Name', 'Title', 'Email', 'Status', 'Department', 'Team', 'Manager', 'Hired On', 'Salary']
        writer = csv.DictWriter(string_output, fieldnames=fieldnames)

        writer.writeheader()
        for node in org_tree:
            emp = node['employee']
            level = node['level']
            emp_dict = self._employee_to_dict(emp)
            emp_dict['Level'] = level
            writer.writerow(emp_dict)

        # Encode and return
        output = BytesIO()
        output.write(string_output.getvalue().encode('utf-8'))
        output.seek(0)

        return output

    def export_org_chart_excel(self, filters: ExportFilterSchema) -> BytesIO:
        """
        Export organizational chart as Excel with visual hierarchy.

        Args:
            filters: Export filter criteria

        Returns:
            BytesIO buffer containing Excel data
        """
        employees = self._apply_filters(filters)
        org_tree = self._build_org_tree(employees)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Org Chart"

        # Define headers
        headers = ['Name', 'Title', 'Email', 'Status', 'Department', 'Team', 'Manager', 'Hired On', 'Salary']

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data with indentation
        for row_num, node in enumerate(org_tree, 2):
            emp = node['employee']
            level = node['level']
            emp_dict = self._employee_to_dict(emp)

            for col_num, header in enumerate(headers, 1):
                value = emp_dict[header]

                # Add indentation to name based on level
                if header == 'Name':
                    value = '  ' * level + value

                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Color code by level
                if level == 0:
                    cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                    cell.font = Font(bold=True)
                elif level == 1:
                    cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def export_org_chart_pdf(self, filters: ExportFilterSchema) -> BytesIO:
        """
        Export organizational chart as visual PDF diagram.

        Args:
            filters: Export filter criteria

        Returns:
            BytesIO buffer containing PDF data
        """
        employees = self._apply_filters(filters)
        org_tree = self._build_org_tree(employees)

        # Create PDF with custom drawing
        output = BytesIO()
        c = canvas.Canvas(output, pagesize=letter)
        width, height = letter

        # Configuration
        box_width = 200
        box_height = 60
        vertical_spacing = 30
        horizontal_spacing = 20
        indent_per_level = 40

        # Starting position
        y_position = height - 50

        # Title
        c.setFont("Helvetica-Bold", 16)
        c.setFillColor(colors.HexColor('#366092'))
        c.drawString(50, y_position, "Organizational Chart")
        y_position -= 40

        # Draw each employee
        for node in org_tree:
            emp = node['employee']
            level = node['level']

            # Calculate position
            x_position = 50 + (level * indent_per_level)

            # Check if we need a new page
            if y_position < 100:
                c.showPage()
                y_position = height - 50
                x_position = 50 + (level * indent_per_level)

            # Draw box
            if level == 0:
                c.setFillColor(colors.HexColor('#366092'))
            elif level == 1:
                c.setFillColor(colors.HexColor('#4A90E2'))
            else:
                c.setFillColor(colors.HexColor('#7FB3D5'))

            c.rect(x_position, y_position - box_height, box_width, box_height, fill=1)

            # Draw text
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x_position + 5, y_position - 15, emp.name[:30])

            c.setFont("Helvetica", 8)
            c.drawString(x_position + 5, y_position - 30, emp.title[:35] if emp.title else 'No Title')

            c.setFont("Helvetica", 7)
            dept_text = emp.department.name if emp.department else 'No Dept'
            c.drawString(x_position + 5, y_position - 45, f"Dept: {dept_text[:20]}")

            # Move to next position
            y_position -= (box_height + vertical_spacing)

        # Save PDF
        c.save()
        output.seek(0)

        return output
